import threading
import openpyxl
from webinit_ import initBrowser
from login import login
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from single_page import single_page
from thread_asking import thread_asking
from url.mobilax import mobilax_urls
    

## spliting urls into arrays for threading

url_splits = []
url_splits = []
threads = []
items = []


thread_amount = thread_asking()
category_pages = [line.strip() for line in mobilax_urls]
split_size = len(category_pages) // thread_amount
remainder = len(category_pages) % thread_amount


start = 0
for i in range(thread_amount):
    chunk_size = split_size + (1 if i < remainder else 0)
    url_split = category_pages[start:start+chunk_size]
    url_splits.append(url_split)
    start += chunk_size



init_driver = login(initBrowser(True))
WebDriverWait(init_driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "app-price")))
cookies = init_driver.execute_script("return document.cookie")
init_driver.quit()
for url_array in url_splits :
    thread = threading.Thread(target=single_page, args=(url_array, items, cookies.split(';')[1]))
    threads.append(thread)
    thread.start()
for thread in threads:
    thread.join()









## saving ZONE
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='SKU')
worksheet.cell(row=1, column=2, value='PRICE')
for row_number, row in enumerate(items, start=2):
    worksheet.cell(row=row_number, column=1, value=row[0])
    worksheet.cell(row=row_number, column=2, value=row[1])
workbook.save('_output/mobilax.xlsx')