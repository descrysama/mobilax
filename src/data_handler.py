import os
import configparser
import openpyxl


def add_data(file_path, items):
    
    file_exists = os.path.isfile(file_path)

    if not file_exists:
        workbook = openpyxl.Workbook()
    else :
        workbook = openpyxl.load_workbook(file_path)
    
    worksheet = workbook.active
    last_row = worksheet.max_row + 1
    for row_number, row in enumerate(items, start=last_row):
        worksheet.cell(row=row_number, column=1, value=row[0])
        worksheet.cell(row=row_number, column=2, value=row[1])
    workbook.save(file_path)

    return


def remove_at_start(file_path):
    file_exists = os.path.isfile(file_path)

    if file_exists:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        worksheet.delete_rows(1, worksheet.max_row)
        workbook.save(file_path)