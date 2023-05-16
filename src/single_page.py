
from webinit_ import initBrowser
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sys;
import openpyxl;
import os;


def find_element_with_retry(driver, target, url):
    while True:
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, target)))
            return
        except:
            print("Element not found, retrying...")
            driver.get(url)
            find_element_with_retry(driver, target, url)
            continue

def single_page(thread_id, url_array, results, cookie) :
    if(len(url_array) < 1) : 
        return
    
    
    driver = initBrowser(True)
    driver.get("https://www.mobilax.fr")
    driver.execute_script("document.cookie='" + cookie + "; path=/; domain=www.mobilax.fr; secure=false; sameSite=Strict;'")

    for telephone_index, category_page in enumerate(url_array):
        print("Thread N°" , thread_id + 1, ":" , telephone_index , "/" , len(url_array))
        try:
            total_items = []
            # Naviguer vers la page de la catégorie
            driver.get(category_page)
            ##os.system("cls")
            # Attendre que les produits soient chargés
            find_element_with_retry(driver, "a.app-product-tile", category_page)

            # Récupérer les liens de tous les produits de la catégorie
            product_elements = driver.find_elements(By.CSS_SELECTOR, "a.app-product-tile")
            product_links = [el.get_attribute('href') for el in product_elements]

            for produit_index, product_link in enumerate(product_links):
                print("Thread N°" , thread_id + 1, "-" , "produits :", produit_index , "/" , len(product_links))
                # Naviguer vers la page du produit
                driver.get(product_link)
                item = ["",""]

                find_element_with_retry(driver, "section.name-container p", product_link)
                driver.execute_script("window.stop();")
                reference_element = driver.find_element(By.CSS_SELECTOR, "section.name-container p")

                find_element_with_retry(driver, "div.current-price p.value.app-typo-h4", product_link)
                prix_element = driver.find_element(By.CSS_SELECTOR, "div.current-price p.value.app-typo-h4")
                driver.execute_script("window.stop();")
                reference_element_text = reference_element.text.replace("RÉF: ", "")
                prix_element_text = prix_element.text.replace(" € HT", "")
                
                item[0] = reference_element_text
                item[1] = prix_element_text
                total_items.append(item)
            file_path = 'output.xlsx'
            file_exists = os.path.isfile(file_path)
            if file_exists:
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = openpyxl.Workbook()

            worksheet = workbook.active
            last_row = worksheet.max_row + 1
            for row_number, row in enumerate(total_items, start=last_row):
                worksheet.cell(row=row_number, column=1, value=row[0])
                worksheet.cell(row=row_number, column=2, value=row[1])
            workbook.save('output.xlsx') 

        except:
            print("thread error")
            driver.quit()
            return        
    driver.quit()
